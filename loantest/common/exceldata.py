#-*-coding: UTF-8 -*-
'''
Created on 2019年10月8日
@author: LIJY
'''
from conf import constant
import openpyxl

logger = constant.MyLogger("run_log","logger","log.txt")






class Exceldata():
    def __init__(self,filepath):
        super().__init__()
        self.filepath = filepath
        
        
    # 定义打开xls的函数
    def open_xls(self):
        self.workbook = openpyxl.load_workbook(self.filepath,data_only = True)
        
    
    def excel_save(self):
        # 保存和关闭数据表
        self.workbook.save(self.filepath)
        self.workbook.close()
 
    
    # 定义获取表单的函数
    def get_sheet(self, sheet):
        
        # 先打开数据表
        self.open_xls()
        
        if isinstance(sheet, int):  # 如果传入的是int，返回该索引表单,从0下标开始
            return self.workbook.worksheets[sheet]
        else:
            if sheet in self.workbook.sheetnames:  # 如果传入的是str,按名字取得sheet
                return self.workbook[sheet]
            else:
                return None
        # 结束后关闭数据表
        self.excel_save()
                          
    
    def get_data(self,sheetname):
        
        '''
        输入起始sheetname,行起始号，行结束号，列起始号，列结束号，每一行为dict，对应str类型的表头和内容
        '''
        list_result = []
        sheet = self.get_sheet(sheetname)
        row_start = 2  # 从第二行开始获取测试数据
        row_end = sheet.max_row
        col_end = sheet.max_column

        for row in range(row_start,row_end+1):
            dic_row = {}
            list_row = []
            head = [] #初始化表头
            
            # 读取表头
            for col in range(1,col_end+1):
                if sheet.cell(1,col).value is not None:
                    head.append(sheet.cell(1,col).value)
                                    
            for col in range(1,col_end+1):
                value = sheet.cell(row,col).value
                # 把取出来的str类型的数据，用eval转换类型
                list_row.append(value)
                # 如果读取到一行空行，就不采用
            if list_row.count(None)==len(list_row):
                continue
            else:
                dic_row = dict(zip(head,list_row))
                list_result.append(dic_row)               
        return list_result
    
    def set_cell_value(self,sheetname,row,col,newvalue):
        
        '''
        设置数据表数据
        '''
        sheet = self.workbook.sheetnames(sheetname)
        sheet.cell(row,col).value = newvalue
        # 关闭数据表
        self.excel_save()

if __name__=='__main__':
    import json
    excel = Exceldata('../data/excel_handler_test.xlsx')
    a = excel.get_data('Sheet1')
#     print(a)
#     print(a[0]['DATA'])
#     mm=a[0]['DATA'].replace('*new_phone*','123456789')
#     print(mm)
#     print('========')
#     
#     print(type(a[0]['EXPECTED']))
#     print(a[0]['EXPECTED'])
#     print(eval(a[0]['EXPECTED']))
# #     print(json.loads(a[0]['EXPECTED']))
# #     print(type(json.loads(a[0]['EXPECTED'])))
#     print('========')
    c = a[0]['EXPECTED']
    print(c)
    
    b = json.loads(c)
    print(b)
                 